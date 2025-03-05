# import openpyxl  
# from openpyxl.styles import PatternFill, Font  
# from openpyxl.worksheet.datavalidation import DataValidation  
# from openpyxl.utils import get_column_letter  
# from openpyxl.formatting.rule import FormulaRule  

# # 1. Create a new workbook  
# workbook = openpyxl.Workbook()  
# sheet = workbook.active  # Get the active sheet  

# # 2. Add headers (optional)  
# sheet['A1'] = 'Student Name'  
# sheet['B1'] = 'Attendance'  

# # 3. Add student names (optional)  
# students = ["Alice", "Bob", "Charlie", "David", "Eve"]  
# for i, student in enumerate(students):  
#     sheet[f'A{i+2}'] = student  

# # 4. Define the range for the data validation  
# start_row = 2  
# end_row = 6  
# start_col = 2  # Column B  
# end_col = 2  

# start_col_letter = get_column_letter(start_col)  
# end_col_letter = get_column_letter(end_col)  

# data_validation_range = f'{start_col_letter}{start_row}:{end_col_letter}{end_row}'  

# # 5. Create the Data Validation object  
# dv = DataValidation(type="list", formula1='"Present,Absent"', allow_blank=False)  

# dv.promptTitle = "Attendance Selection"  
# dv.prompt = "Choose 'Present' or 'Absent' from the dropdown list."  

# dv.errorTitle = "Invalid Input"  
# dv.error = "Please select a value from the dropdown list."  
# dv.errorStyle = 'warning'  

# # 6. Add the Data Validation to the sheet  
# sheet.add_data_validation(dv)  
# dv.add(data_validation_range)  


# # 7. Conditional Formatting (Color Based on Value)  
# def apply_conditional_formatting(ws, start_row, end_row, col_letter):  
#     """Applies conditional formatting to a column based on 'Present' or 'Absent'."""  

#     present_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green  
#     absent_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red  

#     for row in range(start_row, end_row + 1):  
#         cell = ws[f'{col_letter}{row}']  
#         #Initial color setup based on default, to avoid needing user input first  
#         if cell.value == "Present":  
#            cell.fill = present_fill  
#         elif cell.value == "Absent":  
#            cell.fill = absent_fill  


#         #This could be changed to cell.value, but it makes it more robust for changes to Present / Absent  
#         cell.style = 'Normal' #Resets style, need to do otherwise formula messes it up  
#         cell.font = Font(color="000000") #Optional, ensures font remains black  

#         # Create the FormulaRule objects  
#         formula_present = f'=$B${row}="Present"'  
#         rule_present = FormulaRule(formula=[formula_present], fill=present_fill)  

#         formula_absent = f'=$B${row}="Absent"'  
#         rule_absent = FormulaRule(formula=[formula_absent], fill=absent_fill)  


#         conditional_formatting = ws.conditional_formatting  # Corrected variable name  
#         conditional_formatting.add(f'{col_letter}{row}', rule_present)  
#         conditional_formatting.add(f'{col_letter}{row}', rule_absent)  


# # 8. Apply conditional formatting to the attendance column  
# apply_conditional_formatting(sheet, start_row, end_row, start_col_letter)  # Apply to column B  

# # 9. Save the workbook  
# workbook.save("attendance_sheet_colored.xlsx")  

# print("Excel file 'attendance_sheet_colored.xlsx' created successfully!")  


import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def generate_attendance_sheet():
    # Generate date range
    start_date = datetime(2025, 1, 16)
    end_date = datetime(2025, 3, 5)
    
    # Create list of dates
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Create attendance status (with more varied probability)
    np.random.seed(42)  # for reproducibility
    weights = [0.8, 0.2]  # 80% chance of present, 20% chance of absent
    status = np.random.choice(['Present', 'Absent'], size=len(dates), p=weights)
    
    # Create DataFrame
    df = pd.DataFrame({
        'Date': dates,
        'Status': status
    })
    
    # Convert Date to string for better Excel formatting
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
    
    # Create Excel writer
    excel_path = 'Attendance_Sheet_2025.xlsx'
    
    # Write to Excel with styling
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write the dataframe
        df.to_excel(writer, index=False, sheet_name='Attendance')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance']
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        
        # Add title
        worksheet.insert_rows(1)
        worksheet['A1'] = 'Attendance Sheet (01/16/2025 - 03/05/2025)'
        title_cell = worksheet['A1']
        title_cell.font = Font(bold=True, size=16)
        worksheet.merge_cells('A1:B1')
        
        # Style header row
        header_cells = ['A2', 'B2']
        for cell_ref in header_cells:
            cell = worksheet[cell_ref]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Conditional formatting for Status
        for row in range(3, len(df) + 3):
            status_cell = worksheet.cell(row=row, column=2)
            if status_cell.value == 'Present':
                status_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        
        # Add summary statistics
        total_days = len(df)
        present_days = len(df[df['Status'] == 'Present'])
        absent_days = len(df[df['Status'] == 'Absent'])
        
        worksheet.cell(row=len(df) + 4, column=1, value='Total Days')
        worksheet.cell(row=len(df) + 4, column=2, value=total_days)
        worksheet.cell(row=len(df) + 5, column=1, value='Total Present')
        worksheet.cell(row=len(df) + 5, column=2, value=present_days)
        worksheet.cell(row=len(df) + 6, column=1, value='Total Absent')
        worksheet.cell(row=len(df) + 6, column=2, value=absent_days)
    
    print(f"Attendance sheet generated: {excel_path}")
    
    return df

# Generate the attendance sheet
attendance_df = generate_attendance_sheet()
print(attendance_df)